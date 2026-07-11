# -*- coding: utf-8 -*-
"""
把 seen_listings.json 匯出成可讀的 Excel 清單。
每日排程跑完抓取後自動執行，產生 listings.xlsx 並存回 repo。
"""
import json
import os
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import CITIES

STATE_FILE = "seen_listings.json"
OUTPUT_FILE = "listings.xlsx"

# 台灣時區
TW = timezone(timedelta(hours=8))


def build_region_lookup():
    """行政區 -> 分組名稱 的反查表（跨所有城市）。"""
    lookup = {}
    for city, cfg in CITIES.items():
        for group, regions in cfg.get("region_groups", {}).items():
            for r in regions:
                lookup[(city, r)] = group
    return lookup


def main():
    if not os.path.exists(STATE_FILE):
        print("找不到 seen_listings.json，略過匯出")
        return

    with open(STATE_FILE, encoding="utf-8") as f:
        seen = json.load(f)

    if not seen:
        print("清單為空，略過匯出")
        return

    region_lookup = build_region_lookup()

    rows = []
    for uid, item in seen.items():
        city = item.get("city", "")
        region = item.get("region", "")
        group = region_lookup.get((city, region), "其他")
        rows.append({
            "城市": city.replace("_青埔", "（青埔）"),
            "區域分組": group,
            "行政區": region,
            "案名": item.get("name", ""),
            "單價": item.get("price", ""),
            "坪數房型": item.get("address", ""),
            "狀態": item.get("status", ""),
            "首次發現": item.get("first_seen", ""),
        })

    # 排序：城市 -> 區域分組 -> 行政區 -> 案名，讓同區聚在一起
    rows.sort(key=lambda r: (r["城市"], r["區域分組"], r["行政區"], r["案名"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "預售案清單"

    headers = ["城市", "區域分組", "行政區", "案名", "單價", "坪數房型", "狀態", "首次發現"]

    # 標題列
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # 資料列
    body_font = Font(name="Arial")
    for i, r in enumerate(rows, start=2):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=i, column=col, value=r[h])
            c.font = body_font

    # 欄寬
    widths = {"城市": 14, "區域分組": 18, "行政區": 10, "案名": 26,
              "單價": 16, "坪數房型": 20, "狀態": 10, "首次發現": 12}
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths[h]

    # 凍結標題列 + 自動篩選
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # 更新時間註記（放在資料下方，不干擾篩選）
    note_row = len(rows) + 3
    note = f"最後更新：{datetime.now(TW).strftime('%Y-%m-%d %H:%M')}（台灣時間）  共 {len(rows)} 筆　資料來源：樂屋網"
    nc = ws.cell(row=note_row, column=1, value=note)
    nc.font = Font(name="Arial", italic=True, size=9, color="808080")

    wb.save(OUTPUT_FILE)
    print(f"已匯出 {OUTPUT_FILE}：{len(rows)} 筆")


if __name__ == "__main__":
    main()
